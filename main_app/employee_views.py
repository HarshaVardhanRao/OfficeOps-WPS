import json
import math
from datetime import datetime

from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.shortcuts import (HttpResponseRedirect, get_object_or_404,
                              redirect, render)
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from .models import Examination, ExaminationSubject, StudentProfile, StudentMark, Standard, Section, Subject
from django.contrib.auth.decorators import login_required

from .forms import *
from .models import *


def employee_home(request):
    employee = get_object_or_404(Employee, admin=request.user)
    total_Section = Section.objects.all().count()
    total_attendance = AttendanceReport.objects.filter(employee=employee).count()
    total_present = AttendanceReport.objects.filter(employee=employee, status=True).count()
    total_absent = AttendanceReport.objects.filter(employee=employee, status=False).count()
    Sections = Section.objects.all()

    context = {
        'total_attendance': total_attendance,
        'percent_present': 50,
        'percent_absent': 50,
        'total_Section': total_Section,
        'Sections': Sections,
        'data_present': 90,
        'data_absent': 10,
        'data_name': Sections,
        'page_title': 'Employee Homepage'

    }
    return render(request, 'employee_template/home_content.html', context)


@ csrf_exempt
def employee_view_attendance(request):
    employee = get_object_or_404(Employee, admin=request.user)
    if request.method != 'POST':
        Standard = get_object_or_404(Standard, id=employee.Standard.id)
        context = {
            'Sections': Section.objects.filter(Standard=Standard),
            'page_title': 'View Attendance'
        }
        return render(request, 'employee_template/employee_view_attendance.html', context)
    else:
        Section_id = request.POST.get('Section')
        start = request.POST.get('start_date')
       
        try:
            Sec = get_object_or_404(Section, id=Section_id)
            start_date = datetime.strptime(start, "%Y-%m-%d")
            end_date = datetime.strptime(end, "%Y-%m-%d")
            attendance = Attendance.objects.filter(
                date__range=(start_date, end_date), Section=Sec)
            attendance_reports = AttendanceReport.objects.filter(
                attendance__in=attendance, employee=employee)
            json_data = []
            for report in attendance_reports:
                data = {
                    "date":  str(report.attendance.date),
                    "status": report.status
                }
                json_data.append(data)
            return JsonResponse(json.dumps(json_data), safe=False)
        except Exception as e:
            return None

######################Student Attendance#########################
from django.contrib.auth.decorators import login_required
def manager_take_attendance(request):
    employee = get_object_or_404(Employee, admin=request.user)
    Sections = Section.objects.all()
    print(Sections)
    context = {
        'Sections': Sections,
        'page_title': 'Take Attendance'
    }

    return render(request, 'employee_template/employee_take_attendance.html', context)

from django.contrib.auth.decorators import login_required

@login_required
@csrf_exempt
def get_students(request):
    if request.method == "POST":
        section_id = request.POST.get("section_id")

        if not section_id:
            return JsonResponse({"error": "Missing section ID"}, status=400)

        students = StudentProfile.objects.filter(section_id=section_id).values("id", "student")

        return JsonResponse(list(students), safe=False)

    return JsonResponse({"error": "Invalid request method"}, status=400)



@csrf_exempt
def save_student_attendance(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            attendance_date = data.get("date")
            section_id = data.get("section_id")
            student_ids = data.get("student_ids", [])

            sec = Section.objects.get(id = section_id)

            if not attendance_date or not section_id or not student_ids:
                return JsonResponse({"error": "Missing required fields"}, status=400)

            # Get or create an Attendance entry for the section & date
            attendance_entry, created = Attendance.objects.get_or_create(
                section=sec, date=attendance_date
            )

            for student in student_ids:
                student_id = student.get("id")
                status = student.get("status")

                if student_id is not None:
                    # Save attendance record for the student
                    AttendanceReportStudent.objects.update_or_create(
                        attendance=attendance_entry,
                        student_id=student_id,
                        defaults={"status": status}  # 1 = Present, 0 = Absent
                    )

            return JsonResponse({"message": "Attendance saved successfully!"}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)

    return JsonResponse({"error": "Invalid request"}, status=400)


from django.http import JsonResponse
from .models import Attendance

from django.contrib.auth.decorators import login_required

@login_required
def get_attendance_date(request):
    section_id = request.GET.get("section_id")
    if not section_id:
        return JsonResponse({"error": "Invalid section ID"}, status=400)

    attendance_dates = Attendance.objects.filter(section_id=section_id).values("id", "date")
    return JsonResponse(list(attendance_dates), safe=False)



from django.http import JsonResponse
from .models import AttendanceReportStudent

def get_attendance_report(request):
    employee = get_object_or_404(Employee, admin=request.user)
    attendance_id = request.GET.get("attendance_id")
    reports = AttendanceReportStudent.objects.filter(attendance_id=attendance_id).select_related("student")

    data = [
        {"student_name": report.student.student, "status": report.status}
        for report in reports
    ]
    return JsonResponse(data, safe=False)


def manager_update_attendance(request):
    # try:
    #     manager = get_object_or_404(Manager, admin=request.user)
    #     if manager is None:
    #         if request.user.is_superuser:
    #             manager = request.user
    # except:
    #     print("Manager Not found")
    Sections = Section.objects.all()
    context = {
        'Sections': Sections,
        'page_title': 'Update Attendance'
    }

    return render(request, 'employee_template/employee_update_attendance.html', context)

from django.http import JsonResponse
from .models import Attendance
from .models import AttendanceReportStudent

def get_attendance_report(request):
    attendance_id = request.GET.get("attendance_id")
    reports = AttendanceReportStudent.objects.filter(attendance_id=attendance_id).select_related("student")

    data = [
        {"student_name": report.student.student, "status": report.status}
        for report in reports
    ]
    return JsonResponse(data, safe=False)

##############################################################################################################
def employee_apply_leave(request):
    form = LeaveReportEmployeeForm(request.POST or None)
    employee = get_object_or_404(Employee, admin_id=request.user.id)
    context = {
        'form': form,
        'leave_history': LeaveReportEmployee.objects.filter(employee=employee),
        'page_title': 'Apply for leave'
    }
    if request.method == 'POST':
        if form.is_valid():
            try:
                obj = form.save(commit=False)
                obj.employee = employee
                obj.save()
                messages.success(
                    request, "Application for leave has been submitted for review")
                return redirect(reverse('employee_apply_leave'))
            except Exception:
                messages.error(request, "Could not submit")
        else:
            messages.error(request, "Form has errors!")
    return render(request, "employee_template/employee_apply_leave.html", context)


def employee_feedback(request):
    form = FeedbackEmployeeForm(request.POST or None)
    employee = get_object_or_404(Employee, admin_id=request.user.id)
    context = {
        'form': form,
        'feedbacks': FeedbackEmployee.objects.filter(employee=employee),
        'page_title': 'Employee Feedback'

    }
    if request.method == 'POST':
        if form.is_valid():
            try:
                obj = form.save(commit=False)
                obj.employee = employee
                obj.save()
                messages.success(
                    request, "Feedback submitted for review")
                return redirect(reverse('employee_feedback'))
            except Exception:
                messages.error(request, "Could not Submit!")
        else:
            messages.error(request, "Form has errors!")
    return render(request, "employee_template/employee_feedback.html", context)


def employee_view_profile(request):
    employee = get_object_or_404(Employee, admin=request.user)
    form = EmployeeEditForm(request.POST or None, request.FILES or None,
                           instance=employee)
    context = {'form': form,
               'page_title': 'View/Edit Profile'
               }
    if request.method == 'POST':
        try:
            if form.is_valid():
                first_name = form.cleaned_data.get('first_name')
                last_name = form.cleaned_data.get('last_name')
                password = form.cleaned_data.get('password') or None
                address = form.cleaned_data.get('address')
                gender = form.cleaned_data.get('gender')
                passport = request.FILES.get('profile_pic') or None
                admin = employee.admin
                if password != None:
                    admin.set_password(password)
                if passport != None:
                    fs = FileSystemStorage()
                    filename = fs.save(passport.name, passport)
                    passport_url = fs.url(filename)
                    admin.profile_pic = passport_url
                admin.first_name = first_name
                admin.last_name = last_name
                admin.address = address
                admin.gender = gender
                admin.save()
                employee.save()
                messages.success(request, "Profile Updated!")
                return redirect(reverse('employee_view_profile'))
            else:
                messages.error(request, "Invalid Data Provided")
        except Exception as e:
            messages.error(request, "Error Occured While Updating Profile " + str(e))

    return render(request, "employee_template/employee_view_profile.html", context)


@csrf_exempt
def employee_fcmtoken(request):
    token = request.POST.get('token')
    employee_user = get_object_or_404(CustomUser, id=request.user.id)
    try:
        employee_user.fcm_token = token
        employee_user.save()
        return HttpResponse("True")
    except Exception as e:
        return HttpResponse("False")


def employee_view_notification(request):
    employee = get_object_or_404(Employee, admin=request.user)
    notifications = NotificationEmployee.objects.filter(employee=employee)
    context = {
        'notifications': notifications,
        'page_title': "View Notifications"
    }
    return render(request, "employee_template/employee_view_notification.html", context)


def employee_view_salary(request):
    employee = get_object_or_404(Employee, admin=request.user)
    salarys = EmployeeSalary.objects.filter(employee=employee)
    context = {
        'salarys': salarys,
        'page_title': "View Salary"
    }
    return render(request, "employee_template/employee_view_salary.html", context)


@login_required
def employee_download_marks_template(request, exam_id):
    exam = Examination.objects.get(id=exam_id)
    exam_subjects = ExaminationSubject.objects.filter(examination=exam)
    students = StudentProfile.objects.filter(standard=exam.standard, section=exam.section)
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Marks Entry'
    ws.append(['Student Name', 'Student ID'] + [f'{es.subject.name}' for es in exam_subjects])
    for student in students:
        ws.append([student.student, student.id] + ['' for _ in exam_subjects])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=marks_template_{exam.name}.xlsx'
    wb.save(response)
    return response

@csrf_exempt
def employee_upload_marks(request, exam_id):
    exam = Examination.objects.get(id=exam_id)
    exam_subjects = list(ExaminationSubject.objects.filter(examination=exam))
    if request.method == 'POST' and request.FILES.get('file'):
        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            student_id = row[1]
            student = StudentProfile.objects.get(id=student_id)
            for idx, exam_subject in enumerate(exam_subjects):
                marks = row[2+idx]
                if marks is not None and marks != '':
                    StudentMark.objects.update_or_create(
                        student=student,
                        exam_subject=exam_subject,
                        defaults={'marks_obtained': marks}
                    )
        messages.success(request, 'Marks uploaded successfully!')
        return redirect('employee_view_profile')
    return render(request, 'employee_template/employee_upload_marks.html', {'exam': exam})

from django.shortcuts import render, redirect

@login_required
def employee_select_exam_upload(request):
    examinations = Examination.objects.all()
    exam_id = request.GET.get('exam_id')
    if exam_id:
        return render(request, 'employee_template/employee_select_exam_upload.html', {'examinations': examinations, 'exam_id': exam_id})
    return render(request, 'employee_template/employee_select_exam_upload.html', {'examinations': examinations})

from django.shortcuts import render, redirect

@login_required
def employee_add_examination(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        standard_id = request.POST.get('standard')
        section_id = request.POST.get('section')
        subjects = request.POST.getlist('subjects[]')
        max_marks = request.POST.getlist('max_marks[]')
        conducted_dates = request.POST.getlist('conducted_dates[]')
        standard = Standard.objects.get(id=standard_id)
        section = Section.objects.get(id=section_id)
        exam = Examination.objects.create(name=name, standard=standard, section=section)
        for i, subject_name in enumerate(subjects):
            subject_name_clean = subject_name.strip()
            subject = Subject.objects.filter(name__iexact=subject_name_clean).first()
            if not subject:
                subject = Subject.objects.create(name=subject_name_clean)
            ExaminationSubject.objects.create(
                examination=exam,
                subject=subject,
                max_marks=max_marks[i],
                conducted_date=conducted_dates[i]
            )
        return render(request, 'employee_template/employee_add_examination.html', {
            'standards': Standard.objects.all(),
            'sections': Section.objects.all(),
            'exam_id': exam.id
        })
    standards = Standard.objects.all()
    sections = Section.objects.all()
    return render(request, 'employee_template/employee_add_examination.html', {
        'standards': standards,
        'sections': sections
    })

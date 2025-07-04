import json

from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.shortcuts import (HttpResponseRedirect, get_object_or_404,redirect, render)
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
import json
from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Attendance, StudentProfile
from .forms import *
from .models import *
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect
from django.urls import reverse
from .models import Examination, ExaminationSubject, Subject, Standard, Section, StudentProfile, StudentMark
from .forms import *
import openpyxl
from django.http import HttpResponse

def manager_home(request):
    manager = get_object_or_404(Manager, admin=request.user)
    total_employees = Employee.objects.all().count()
    total_leave = LeaveReportManager.objects.filter(manager=manager).count()
    Sections = Section.objects.all()
    total_Section = Sections.count()
    attendance_list = Attendance.objects.filter(section__in=Sections)
    total_attendance = attendance_list.count()
    attendance_list = []
    Section_list = []
    for sec in Sections:
        attendance_count = Attendance.objects.filter(section=sec).count()
        Section_list.append(sec.name)
        attendance_list.append(attendance_count)
    context = {
        'page_title': 'Manager Panel - ' + str(manager.admin.last_name),
        'total_employees': total_employees,
        'total_attendance': total_attendance,
        'total_leave': total_leave,
        'total_Section': total_Section,
        'Section_list': Section_list,
        'attendance_list': attendance_list
    }
    return render(request, 'manager_template/home_content.html', context)


def manager_take_attendance(request):
    # manager = get_object_or_404(Manager, admin=request.user)
    Sections = Section.objects.all()
    print(Sections)
    context = {
        'Sections': Sections,
        'page_title': 'Take Attendance'
    }

    return render(request, 'manager_template/manager_take_attendance.html', context)




######################## Student Attendance Rechecked ####################################################

@csrf_exempt
def get_students(request):
    if request.method == "POST":
        section_id = request.POST.get("section_id")
        students = StudentProfile.objects.filter(section_id=section_id).values("id", "student")

        return JsonResponse(list(students), safe=False)

    return JsonResponse({"error": "Invalid request"}, status=400)

@csrf_exempt
def save_student_attendance(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            attendance_date = data.get("date")
            section_id = data.get("section_id")
            student_ids = data.get("student_ids", [])

            sec = Section.objects.get(id = section_id)

            if not attendance_date or not section_id or not student_ids:
                return JsonResponse({"error": "Missing required fields"}, status=400)

            # Get or create an Attendance entry for the section & date
            attendance_entry, created = Attendance.objects.get_or_create(
                section=sec, date=attendance_date
            )

            for student in student_ids:
                student_id = student.get("id")
                status = student.get("status")

                if student_id is not None:
                    # Save attendance record for the student
                    AttendanceReportStudent.objects.update_or_create(
                        attendance=attendance_entry,
                        student_id=student_id,
                        defaults={"status": status}  # 1 = Present, 0 = Absent
                    )

            return JsonResponse({"message": "Attendance saved successfully!"}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)

    return JsonResponse({"error": "Invalid request"}, status=400)


from django.http import JsonResponse
from .models import Attendance
def get_attendance_dates(request):
    section_id = request.GET.get("section_id")
    if not section_id:
        return JsonResponse({"error": "Invalid section ID"}, status=400)

    attendance_dates = Attendance.objects.filter(section_id=section_id).values("id", "date")
    return JsonResponse(list(attendance_dates), safe=False)


from django.http import JsonResponse
from .models import AttendanceReportStudent

def get_attendance_report(request):
    attendance_id = request.GET.get("attendance_id")
    reports = AttendanceReportStudent.objects.filter(attendance_id=attendance_id).select_related("student")

    data = [
        {"student_name": report.student.student, "status": report.status}
        for report in reports
    ]
    return JsonResponse(data, safe=False)
################################  employee attandance ###############################################
from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import Employee, AttendanceEmployee, AttendanceReport

def take_employee_attendance(request):
    return render(request, "manager_template/take_employee_attendance.html")

@csrf_exempt
def get_employees(request):
    if request.method == "POST":
        data = request.POST
        employees = Employee.objects.all().values("id", "admin__first_name", "admin__last_name")

        employee_list = [
            {"id": emp["id"], "name": f"{emp['admin__first_name']} {emp['admin__last_name']}"}
            for emp in employees
        ]
        return JsonResponse(employee_list, safe=False)

@csrf_exempt
def save_employee_attendance(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            date = data.get("date")
            employees = data.get("employees")

            # Create an attendance entry for the date
            attendance_entry, created = AttendanceEmployee.objects.get_or_create(date=date)
            print(attendance_entry)

            # Save attendance for each employee
            for emp in employees:
                employee_obj = Employee.objects.get(id=emp["id"])
                print(employee_obj)
                AttendanceReport.objects.update_or_create(
                    employee=employee_obj,
                    attendance=attendance_entry,
                    defaults={"status": emp["status"]}
                )

            return JsonResponse({"message": "Attendance saved successfully"}, status=200)
        except Exception as e:
            return JsonResponse({"message": "Error: " + str(e)}, status=500)


###############################################################################
@csrf_exempt
def get_employees(request):
    Section_id = request.POST.get('section')
    print(Section_id)
    try:
        section = get_object_or_404(Section, id=Section_id)
        employees = Employee.objects.filter(Standard_id=section.standard.id)
        employee_data = []
        for employee in employees:
            data = {
                "id": employee.id,
                "name": employee.admin.last_name + " " + employee.admin.first_name
            }
            employee_data.append(data)
        return JsonResponse(json.dumps(employee_data), content_type='application/json', safe=False)
    except Exception as e:
        return e


import json

from django.db import transaction
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Attendance, AttendanceReportStudent, Section, StudentProfile
import json
from django.http import JsonResponse



@csrf_exempt
def save_attendance(request):
    employee_data = request.POST.get('employee_ids')
    date = request.POST.get('date')
    Section_id = request.POST.get('section')
    employees = json.loads(employee_data)
    try:
        Sec = get_object_or_404(Section, id=Section_id)

        # Check if an attendance object already exists for the given date
        attendance, created = Attendance.objects.get_or_create(section=Sec, date=date)

        for employee_dict in employees:
            employee = get_object_or_404(Employee, id=employee_dict.get('id'))

            # Check if an attendance report already exists for the employee and the attendance object
            attendance_report, report_created = AttendanceReport.objects.get_or_create(employee=employee, attendance=attendance)

            # Update the status only if the attendance report was newly created
            if report_created:
                attendance_report.status = employee_dict.get('status')
                attendance_report.save()

    except Exception as e:
        return None

    return HttpResponse("OK")


def manager_update_attendance(request):
    try:
        manager = get_object_or_404(Manager, admin=request.user)
        if manager is None:
            if request.user.is_superuser:
                manager = request.user
    except:
        print("Manager Not found")
    Sections = Section.objects.all()
    context = {
        'Sections': Sections,
        'page_title': 'Update Attendance'
    }

    return render(request, 'manager_template/manager_update_attendance.html', context)


@csrf_exempt
def get_employee_attendance(request):
    attendance_date_id = request.POST.get('attendance_date_id')
    try:
        date = get_object_or_404(Attendance, id=attendance_date_id)
        attendance_data = AttendanceReport.objects.filter(attendance=date)
        employee_data = []
        for attendance in attendance_data:
            data = {"id": attendance.employee.admin.id,
                    "name": attendance.employee.admin.last_name + " " + attendance.employee.admin.first_name,
                    "status": attendance.status}
            employee_data.append(data)
        return JsonResponse(json.dumps(employee_data), content_type='application/json', safe=False)
    except Exception as e:
        return e


@csrf_exempt
def update_attendance(request):
    employee_data = request.POST.get('employee_ids')
    date = request.POST.get('date')
    employees = json.loads(employee_data)
    try:
        attendance = get_object_or_404(Attendance, id=date)

        for employee_dict in employees:
            employee = get_object_or_404(
                Employee, admin_id=employee_dict.get('id'))
            attendance_report = get_object_or_404(AttendanceReport, employee=employee, attendance=attendance)
            attendance_report.status = employee_dict.get('status')
            attendance_report.save()
    except Exception as e:
        return None

    return HttpResponse("OK")


def manager_apply_leave(request):
    form = LeaveReportManagerForm(request.POST or None)
    manager = get_object_or_404(Manager, admin_id=request.user.id)
    context = {
        'form': form,
        'leave_history': LeaveReportManager.objects.filter(manager=manager),
        'page_title': 'Apply for Leave'
    }
    if request.method == 'POST':
        if form.is_valid():
            try:
                obj = form.save(commit=False)
                obj.manager = manager
                obj.save()
                messages.success(
                    request, "Application for leave has been submitted for review")
                return redirect(reverse('manager_apply_leave'))
            except Exception:
                messages.error(request, "Could not apply!")
        else:
            messages.error(request, "Form has errors!")
    return render(request, "manager_template/manager_apply_leave.html", context)


def manager_feedback(request):
    form = FeedbackManagerForm(request.POST or None)
    manager = get_object_or_404(Manager, admin_id=request.user.id)
    context = {
        'form': form,
        'feedbacks': FeedbackManager.objects.filter(manager=manager),
        'page_title': 'Add Feedback'
    }
    if request.method == 'POST':
        if form.is_valid():
            try:
                obj = form.save(commit=False)
                obj.manager = manager
                obj.save()
                messages.success(request, "Feedback submitted for review")
                return redirect(reverse('manager_feedback'))
            except Exception:
                messages.error(request, "Could not Submit!")
        else:
            messages.error(request, "Form has errors!")
    return render(request, "manager_template/manager_feedback.html", context)


def manager_view_profile(request):
    manager = get_object_or_404(Manager, admin=request.user)
    form = ManagerEditForm(request.POST or None, request.FILES or None,instance=manager)
    context = {'form': form, 'page_title': 'View/Update Profile'}
    if request.method == 'POST':
        try:
            if form.is_valid():
                first_name = form.cleaned_data.get('first_name')
                last_name = form.cleaned_data.get('last_name')
                password = form.cleaned_data.get('password') or None
                address = form.cleaned_data.get('address')
                gender = form.cleaned_data.get('gender')
                passport = request.FILES.get('profile_pic') or None
                admin = manager.admin
                if password != None:
                    admin.set_password(password)
                if passport != None:
                    fs = FileSystemStorage()
                    filename = fs.save(passport.name, passport)
                    passport_url = fs.url(filename)
                    admin.profile_pic = passport_url
                admin.first_name = first_name
                admin.last_name = last_name
                admin.address = address
                admin.gender = gender
                admin.save()
                manager.save()
                messages.success(request, "Profile Updated!")
                return redirect(reverse('manager_view_profile'))
            else:
                messages.error(request, "Invalid Data Provided")
                return render(request, "manager_template/manager_view_profile.html", context)
        except Exception as e:
            messages.error(
                request, "Error Occured While Updating Profile " + str(e))
            return render(request, "manager_template/manager_view_profile.html", context)

    return render(request, "manager_template/manager_view_profile.html", context)


@csrf_exempt
def manager_fcmtoken(request):
    token = request.POST.get('token')
    try:
        manager_user = get_object_or_404(CustomUser, id=request.user.id)
        manager_user.fcm_token = token
        manager_user.save()
        return HttpResponse("True")
    except Exception as e:
        return HttpResponse("False")


def manager_view_notification(request):
    manager = get_object_or_404(Manager, admin=request.user)
    notifications = NotificationManager.objects.filter(manager=manager)
    context = {
        'notifications': notifications,
        'page_title': "View Notifications"
    }
    return render(request, "manager_template/manager_view_notification.html", context)


def manager_add_examination(request):
    if request.method == 'POST':
        # You should create a form for this, here is a simple version
        name = request.POST.get('name')
        standard_id = request.POST.get('standard')
        section_id = request.POST.get('section')
        subjects = request.POST.getlist('subjects[]')
        max_marks = request.POST.getlist('max_marks[]')
        conducted_dates = request.POST.getlist('conducted_dates[]')
        standard = Standard.objects.get(id=standard_id)
        section = Section.objects.get(id=section_id)
        exam = Examination.objects.create(name=name, standard=standard, section=section)
        for i, subject_name in enumerate(subjects):
            subject_name_clean = subject_name.strip()
            subject = Subject.objects.filter(name__iexact=subject_name_clean).first()
            if not subject:
                subject = Subject.objects.create(name=subject_name_clean)
            ExaminationSubject.objects.create(
                examination=exam,
                subject=subject,
                max_marks=max_marks[i],
                conducted_date=conducted_dates[i]
            )
        return redirect('manager_download_marks_template', exam_id=exam.id)
    standards = Standard.objects.all()
    sections = Section.objects.all()
    return render(request, 'manager_template/manager_add_examination.html', {
        'standards': standards,
        'sections': sections
    })

def manager_download_marks_template(request, exam_id):
    exam = Examination.objects.get(id=exam_id)
    exam_subjects = ExaminationSubject.objects.filter(examination=exam)
    students = StudentProfile.objects.filter(standard=exam.standard, section=exam.section)
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Marks Entry'
    ws.append(['Student Name', 'Student ID'] + [f'{es.subject.name}' for es in exam_subjects])
    for student in students:
        ws.append([student.student, student.id] + ['' for _ in exam_subjects])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=marks_template_{exam.name}.xlsx'
    wb.save(response)
    return response

@csrf_exempt
def manager_upload_marks(request, exam_id):
    exam = Examination.objects.get(id=exam_id)
    exam_subjects = list(ExaminationSubject.objects.filter(examination=exam))
    if request.method == 'POST' and request.FILES.get('file'):
        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            student_id = row[1]
            student = StudentProfile.objects.get(id=student_id)
            for idx, exam_subject in enumerate(exam_subjects):
                marks = row[2+idx]
                if marks is not None and marks != '':
                    StudentMark.objects.update_or_create(
                        student=student,
                        exam_subject=exam_subject,
                        defaults={'marks_obtained': marks}
                    )
        return redirect('manager_home')
    return render(request, 'manager_template/manager_upload_marks.html', {'exam': exam})
